'''
对 RR 神经元构建图，进行网络分析
'''
import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import h5py
import igraph as ig
import seaborn as sns
import pandas as pd
from scipy import sparse
import warnings
import logging
from scipy.stats import zscore
from openpyxl import load_workbook, Workbook
from scipy.signal import butter, filtfilt
import scipy.io

warnings.filterwarnings('ignore')

# matplotlib 字体设置
plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial']
plt.rcParams['axes.unicode_minus'] = False

# ==================== 分析模式配置 ====================
# 设置为 True 仅使用兴奋性RR神经元，False 使用全部RR神经元
USE_EXCITATORY_ONLY = True  # 默认使用全部RR神经元

# 选择分析的小鼠：'m79' 或 'm21'
MOUSE_ID = 'm79'

# ==================== 配置参数 ====================
# 根据选择的小鼠设置数据路径
DATA_FILE = f"C:\\Users\\wangy\\Desktop\\IC\\{MOUSE_ID}\\wholebrain_output.mat"
RR_INDICES_CSV = f"C:\\Users\\wangy\\Desktop\\IC\\{MOUSE_ID}\\rr_neuron_original_indices.csv"

# 滤波参数
SAMPLING_RATE = 4.0  # Hz
HIGH_PASS_CUTOFF = 0.05  # Hz

# 阈值扫描参数
SCAN_THRESHOLDS = np.arange(0.1, 0.5, 0.05)  # 0.1到0.5，步长0.05

# Hub节点判断标准 - 统一使用z-score > 1.5
HUB_ZSCORE_THRESHOLD = 1.5

# 简化的日志设置 - 只显示信息内容
logging.basicConfig(level=logging.INFO,
                    format="%(message)s",  # 只显示信息内容
                    handlers=[logging.StreamHandler(sys.stdout)])
log = logging.getLogger(__name__)


# -------------------- 滤波函数 --------------------
def high_pass_filter(data, cutoff, fs, order=4):
    """应用高通滤波器"""
    nyquist = 0.5 * fs
    normal_cutoff = cutoff / nyquist
    b, a = butter(order, normal_cutoff, btype='high', analog=False)
    filtered_data = filtfilt(b, a, data, axis=1)
    return filtered_data


# -------------------- CSV转Excel功能 (修改为保留三位有效数字) --------------------
def csv_to_excel(csv_path, excel_path=None):
    """将CSV文件转换为格式化的Excel文件，数值保留三位有效数字"""
    if excel_path is None:
        excel_path = csv_path.replace('.csv', '_formatted.xlsx')
    
    try:
        # 读取 CSV
        df = pd.read_csv(csv_path)
        
        # 处理数值列，保留三位有效数字
        for col in df.columns:
            if df[col].dtype in [np.float64, np.float32, np.int64, np.int32]:
                # 对数值列应用格式化，保留三位有效数字
                df[col] = df[col].apply(lambda x: float(f"{x:.3g}") if pd.notnull(x) and isinstance(x, (int, float)) else x)
        
        # 如果 Excel 文件不存在，创建一个
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = "Network Metrics"

        # 清空已有内容（避免重复写入）
        ws.delete_rows(1, ws.max_row)

        # 写入表头
        ws.append(list(df.columns))

        # 写入每一行
        for _, row in df.iterrows():
            # 保证列表类数据以字符串形式写入
            clean_row = [str(x) if isinstance(x, (list, dict)) else x for x in row.values]
            ws.append(clean_row)

        # 自适应列宽
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

        # 保存
        wb.save(excel_path)
        
        log.info(f"✅ CSV转Excel完成！文件已保存到: {excel_path}")
        return excel_path
        
    except Exception as e:
        log.error(f"❌ CSV转Excel时出错: {e}")
        return None


# -------------------- 数据加载 (修改为支持仅兴奋性神经元) --------------------
def load_rr_neurons_data(file_path, rr_indices_csv_path, apply_filter=True):
    """从CSV文件加载RR神经元的荧光信号和坐标数据"""
    rr_df = pd.read_csv(rr_indices_csv_path)
    
    # 根据开关选择神经元类型
    if USE_EXCITATORY_ONLY:
        # 仅使用兴奋性神经元
        rr_df = rr_df[rr_df['category'] == 'exc']
        log.info("🎯 分析模式: 仅使用兴奋性RR神经元")
    else:
        # 使用全部RR神经元
        log.info("🎯 分析模式: 使用全部RR神经元 (兴奋性+抑制性)")
    
    rr_indices = rr_df['neuron_index'].values
    rr_categories = rr_df['category'].values
    
    log.info(f"加载了 {len(rr_indices)} 个 RR 神经元索引")
    log.info(f"兴奋性神经元: {np.sum(rr_categories == 'exc')} 个, 抑制性神经元: {np.sum(rr_categories == 'inh')} 个")

    with h5py.File(file_path, 'r') as f:
        fluorescence = f['whole_trace_ori'][:]
        coordinates = f['whole_center'][:]

    fluorescence = fluorescence.T
    fluorescence_rr = fluorescence[rr_indices, :]

    # 应用高通滤波
    if apply_filter:
        log.info(f"应用高通滤波: 截止频率 {HIGH_PASS_CUTOFF} Hz, 采样率 {SAMPLING_RATE} Hz")
        fluorescence_rr = high_pass_filter(fluorescence_rr, HIGH_PASS_CUTOFF, SAMPLING_RATE)
        log.info("✅ 高通滤波完成")

    coords_used = coordinates[:3, :] if coordinates.shape[0] >= 3 else coordinates
    coordinates_rr = coords_used[:, rr_indices].T

    log.info(f"RR 荧光形状: {fluorescence_rr.shape}, RR 坐标形状: {coordinates_rr.shape}")
    return fluorescence_rr, coordinates_rr, rr_indices, rr_categories


# ========== 从RR分析加载时间段信息 (修改为支持8个时间段) ==========
def load_stimulus_periods_from_rr_analysis(data_path):
    """从RR分析保存的文件中加载8个时间段信息"""
    try:
        # 尝试加载npy文件 - 现在有6个基础时间段
        ic2_file = os.path.join(data_path, "stimulus_periods_ic2.npy")
        ic4_file = os.path.join(data_path, "stimulus_periods_ic4.npy")
        lc2_file = os.path.join(data_path, "stimulus_periods_lc2.npy")
        lc4_file = os.path.join(data_path, "stimulus_periods_lc4.npy")
        baseline_file = os.path.join(data_path, "stimulus_periods_baseline.npy")
        blank_screen_file = os.path.join(data_path, "stimulus_periods_blank_screen.npy")
        
        required_files = [ic2_file, ic4_file, lc2_file, lc4_file, baseline_file, blank_screen_file]
        
        if all(os.path.exists(f) for f in required_files):
            ic2_time_indices = np.load(ic2_file)
            ic4_time_indices = np.load(ic4_file)
            lc2_time_indices = np.load(lc2_file)
            lc4_time_indices = np.load(lc4_file)
            baseline_time_indices = np.load(baseline_file)
            blank_screen_indices = np.load(blank_screen_file)
            
            log.info(f"✅ 从npy文件加载6个基础时间段信息成功")
            
            # 合并IC2和IC4得到IC
            ic_time_indices = np.concatenate([ic2_time_indices, ic4_time_indices])
            # 合并LC2和LC4得到LC
            lc_time_indices = np.concatenate([lc2_time_indices, lc4_time_indices])
            
            # 验证数据有效性
            log.info(f"   基础时间段统计:")
            log.info(f"   IC2: {len(ic2_time_indices)}个时间点")
            log.info(f"   IC4: {len(ic4_time_indices)}个时间点")
            log.info(f"   LC2: {len(lc2_time_indices)}个时间点")
            log.info(f"   LC4: {len(lc4_time_indices)}个时间点")
            log.info(f"   基线: {len(baseline_time_indices)}个时间点")
            log.info(f"   空白屏幕: {len(blank_screen_indices)}个时间点")
            log.info(f"   合并IC: {len(ic_time_indices)}个时间点")
            log.info(f"   合并LC: {len(lc_time_indices)}个时间点")
            
            return (ic_time_indices, lc_time_indices, 
                    ic2_time_indices, ic4_time_indices, 
                    lc2_time_indices, lc4_time_indices, 
                    baseline_time_indices, blank_screen_indices)
        else:
            missing_files = [os.path.basename(f) for f in required_files if not os.path.exists(f)]
            log.warning(f"❌ 未找到完整的时间段npy文件，缺失: {missing_files}")
            
            # 尝试加载旧的四个时间段格式（向后兼容）
            return load_old_four_periods_format(data_path)
        
    except Exception as e:
        log.error(f"❌ 加载时间段信息失败: {e}")
        return load_old_four_periods_format(data_path)


def load_old_four_periods_format(data_path):
    """加载旧的四个时间段格式（向后兼容）"""
    try:
        # 尝试加载旧的四个时间段文件
        ic_file = os.path.join(data_path, "stimulus_periods_ic.npy")
        lc_file = os.path.join(data_path, "stimulus_periods_lc.npy")
        baseline_file = os.path.join(data_path, "stimulus_periods_baseline.npy")
        blank_screen_file = os.path.join(data_path, "stimulus_periods_blank_screen.npy")
        
        if all(os.path.exists(f) for f in [ic_file, lc_file, baseline_file, blank_screen_file]):
            ic_time_indices = np.load(ic_file)
            lc_time_indices = np.load(lc_file)
            baseline_time_indices = np.load(baseline_file)
            blank_screen_indices = np.load(blank_screen_file)
            
            log.info(f"✅ 从旧格式加载4个基础时间段信息成功")
            log.info(f"   IC时间段: {len(ic_time_indices)}个时间点")
            log.info(f"   LC时间段: {len(lc_time_indices)}个时间点")
            log.info(f"   基线时间段: {len(baseline_time_indices)}个时间点")
            log.info(f"   空白屏幕时间段: {len(blank_screen_indices)}个时间点")
            
            # 对于旧格式，将IC作为IC2和IC的合并，LC作为LC2和LC的合并，IC4和LC4为空
            return (ic_time_indices, lc_time_indices, 
                    ic_time_indices, np.array([]), 
                    lc_time_indices, np.array([]), 
                    baseline_time_indices, blank_screen_indices)
        else:
            log.warning("❌ 未找到任何时间段文件，使用备用分割方法")
            return (None, None, None, None, None, None, None, None)
    except Exception as e:
        log.error(f"❌ 加载旧格式时间段信息失败: {e}")
        return (None, None, None, None, None, None, None, None)


# ========== 使用RR分析保存的时间段分割数据 (修改为支持8个时间段) ==========
def split_data_into_eight_periods(fluorescence_rr, data_path):
    """使用RR分析保存的时间段信息分割数据为8个部分"""
    # 尝试从RR分析保存的文件加载时间段信息
    (ic_time_indices, lc_time_indices, 
     ic2_time_indices, ic4_time_indices, 
     lc2_time_indices, lc4_time_indices, 
     baseline_time_indices, blank_screen_indices) = load_stimulus_periods_from_rr_analysis(data_path)
    
    if all(x is not None for x in [ic_time_indices, lc_time_indices, 
                                    ic2_time_indices, ic4_time_indices, 
                                    lc2_time_indices, lc4_time_indices, 
                                    baseline_time_indices, blank_screen_indices]):
        n_timepoints = fluorescence_rr.shape[1]
        
        # 确保时间点不超出数据范围
        ic_time_indices = [int(i) for i in ic_time_indices if i < n_timepoints]
        lc_time_indices = [int(i) for i in lc_time_indices if i < n_timepoints]
        ic2_time_indices = [int(i) for i in ic2_time_indices if i < n_timepoints]
        ic4_time_indices = [int(i) for i in ic4_time_indices if i < n_timepoints]
        lc2_time_indices = [int(i) for i in lc2_time_indices if i < n_timepoints]
        lc4_time_indices = [int(i) for i in lc4_time_indices if i < n_timepoints]
        baseline_time_indices = [int(i) for i in baseline_time_indices if i < n_timepoints]
        blank_screen_indices = [int(i) for i in blank_screen_indices if i < n_timepoints]
        
        ic_data = fluorescence_rr[:, ic_time_indices] if len(ic_time_indices) > 0 else np.empty((fluorescence_rr.shape[0], 0))
        lc_data = fluorescence_rr[:, lc_time_indices] if len(lc_time_indices) > 0 else np.empty((fluorescence_rr.shape[0], 0))
        ic2_data = fluorescence_rr[:, ic2_time_indices] if len(ic2_time_indices) > 0 else np.empty((fluorescence_rr.shape[0], 0))
        ic4_data = fluorescence_rr[:, ic4_time_indices] if len(ic4_time_indices) > 0 else np.empty((fluorescence_rr.shape[0], 0))
        lc2_data = fluorescence_rr[:, lc2_time_indices] if len(lc2_time_indices) > 0 else np.empty((fluorescence_rr.shape[0], 0))
        lc4_data = fluorescence_rr[:, lc4_time_indices] if len(lc4_time_indices) > 0 else np.empty((fluorescence_rr.shape[0], 0))
        baseline_data = fluorescence_rr[:, baseline_time_indices] if len(baseline_time_indices) > 0 else np.empty((fluorescence_rr.shape[0], 0))
        blank_screen_data = fluorescence_rr[:, blank_screen_indices] if len(blank_screen_indices) > 0 else np.empty((fluorescence_rr.shape[0], 0))
        
        log.info(f"✅ 使用RR分析保存的时间段分割数据 (8个时间段):")
        log.info(f"  合并IC: {ic_data.shape[1]}帧 (原始: {len(ic_time_indices)}帧)")
        log.info(f"  合并LC: {lc_data.shape[1]}帧 (原始: {len(lc_time_indices)}帧)")
        log.info(f"  IC2刺激: {ic2_data.shape[1]}帧 (原始: {len(ic2_time_indices)}帧)")
        log.info(f"  IC4刺激: {ic4_data.shape[1]}帧 (原始: {len(ic4_time_indices)}帧)")
        log.info(f"  LC2刺激: {lc2_data.shape[1]}帧 (原始: {len(lc2_time_indices)}帧)")
        log.info(f"  LC4刺激: {lc4_data.shape[1]}帧 (原始: {len(lc4_time_indices)}帧)")
        log.info(f"  基线: {baseline_data.shape[1]}帧 (原始: {len(baseline_time_indices)}帧)")
        log.info(f"  空白屏幕: {blank_screen_data.shape[1]}帧 (原始: {len(blank_screen_indices)}帧)")
        
        return ic_data, lc_data, ic2_data, ic4_data, lc2_data, lc4_data, baseline_data, blank_screen_data
    else:
        # 回退到默认分割方法
        log.warning("❌ 无法加载RR分析的时间段信息，使用默认分割方法")
        return split_data_into_eight_periods_fallback(fluorescence_rr)


def split_data_into_eight_periods_fallback(fluorescence_rr):
    """默认的分割方法（备用）"""
    n_timepoints = fluorescence_rr.shape[1]
    
    if n_timepoints < 8:
        # 如果数据太少，均匀分配
        log.warning("数据点太少，使用均匀分配")
        quarter = max(1, n_timepoints // 4)
        ic2_data = fluorescence_rr[:, :quarter]
        ic4_data = fluorescence_rr[:, quarter:2*quarter] if n_timepoints >= 2*quarter else np.empty((fluorescence_rr.shape[0], 0))
        lc2_data = fluorescence_rr[:, 2*quarter:3*quarter] if n_timepoints >= 3*quarter else np.empty((fluorescence_rr.shape[0], 0))
        lc4_data = fluorescence_rr[:, 3*quarter:] if n_timepoints >= 3*quarter else np.empty((fluorescence_rr.shape[0], 0))
        
        # 合并IC和LC
        ic_data = np.concatenate([ic2_data, ic4_data], axis=1) if ic2_data.shape[1] > 0 or ic4_data.shape[1] > 0 else np.empty((fluorescence_rr.shape[0], 0))
        lc_data = np.concatenate([lc2_data, lc4_data], axis=1) if lc2_data.shape[1] > 0 or lc4_data.shape[1] > 0 else np.empty((fluorescence_rr.shape[0], 0))
        
        # 基线和空白屏幕设为空
        baseline_data = np.empty((fluorescence_rr.shape[0], 0))
        blank_screen_data = np.empty((fluorescence_rr.shape[0], 0))
    else:
        # 八等分分割
        eighth = n_timepoints // 8
        
        ic2_data = fluorescence_rr[:, :eighth]
        ic4_data = fluorescence_rr[:, eighth:2*eighth] if n_timepoints >= 2*eighth else np.empty((fluorescence_rr.shape[0], 0))
        lc2_data = fluorescence_rr[:, 2*eighth:3*eighth] if n_timepoints >= 3*eighth else np.empty((fluorescence_rr.shape[0], 0))
        lc4_data = fluorescence_rr[:, 3*eighth:4*eighth] if n_timepoints >= 4*eighth else np.empty((fluorescence_rr.shape[0], 0))
        baseline_data = fluorescence_rr[:, 4*eighth:5*eighth] if n_timepoints >= 5*eighth else np.empty((fluorescence_rr.shape[0], 0))
        blank_screen_data = fluorescence_rr[:, 5*eighth:6*eighth] if n_timepoints >= 6*eighth else np.empty((fluorescence_rr.shape[0], 0))
        
        # 合并IC和LC
        ic_data = np.concatenate([ic2_data, ic4_data], axis=1) if ic2_data.shape[1] > 0 or ic4_data.shape[1] > 0 else np.empty((fluorescence_rr.shape[0], 0))
        lc_data = np.concatenate([lc2_data, lc4_data], axis=1) if lc2_data.shape[1] > 0 or lc4_data.shape[1] > 0 else np.empty((fluorescence_rr.shape[0], 0))
    
    log.info(f"使用八等分分割:")
    log.info(f"  合并IC: {ic_data.shape[1]}帧 (IC2: {ic2_data.shape[1]}帧, IC4: {ic4_data.shape[1]}帧)")
    log.info(f"  合并LC: {lc_data.shape[1]}帧 (LC2: {lc2_data.shape[1]}帧, LC4: {lc4_data.shape[1]}帧)")
    log.info(f"  基线: {baseline_data.shape[1]}帧")
    log.info(f"  空白屏幕: {blank_screen_data.shape[1]}帧")
    
    return ic_data, lc_data, ic2_data, ic4_data, lc2_data, lc4_data, baseline_data, blank_screen_data


# -------------------- 相关性计算 --------------------
def calculate_correlation_matrix_rr(fluorescence_rr):
    """基于所有时间点计算相关性矩阵"""
    n_neurons = fluorescence_rr.shape[0]
    log.info(f"计算 {n_neurons} 个 RR 神经元的相关性矩阵...")

    means = np.mean(fluorescence_rr, axis=1, keepdims=True)
    stds = np.std(fluorescence_rr, axis=1, keepdims=True)
    stds[stds == 0] = 1.0

    fluorescence_normalized = (fluorescence_rr - means) / stds
    correlation_matrix = np.corrcoef(fluorescence_normalized)
    np.fill_diagonal(correlation_matrix, 0)
    return correlation_matrix


# -------------------- 网络构建与高级分析 --------------------
def create_network_analysis(correlation_matrix, threshold=0.2, rr_indices=None):
    """创建网络并计算多种网络指标"""
    log.info(f"正在创建网络 (阈值={threshold})...")
    mask = np.abs(correlation_matrix) >= threshold
    np.fill_diagonal(mask, False)

    sparse_matrix = sparse.coo_matrix(mask.astype(int))
    sources = sparse_matrix.row
    targets = sparse_matrix.col
    upper_triangular = sources < targets
    sources = sources[upper_triangular]
    targets = targets[upper_triangular]

    g = ig.Graph()
    n_nodes = correlation_matrix.shape[0]
    g.add_vertices(n_nodes)
    if len(sources) > 0:
        edges = list(zip(sources.tolist(), targets.tolist()))
        g.add_edges(edges)

    # 基础网络指标
    degrees = np.array(g.degree())
    clustering_coeff = calculate_clustering_coefficient(g)
    connected_components = g.components()
    largest_component = connected_components.giant() if len(connected_components) > 0 else None
    
    # 计算所有网络指标
    network_metrics = calculate_all_network_metrics(g, degrees, largest_component, rr_indices)
    
    return g, degrees, network_metrics


def calculate_clustering_coefficient(g):
    """计算聚类系数"""
    try:
        return g.transitivity_avglocal()
    except AttributeError:
        try:
            return g.transitivity_undirected()
        except AttributeError:
            return calculate_clustering_manual(g)


def calculate_clustering_manual(g):
    """手动计算聚类系数"""
    clustering_coeffs = []
    for node in range(g.vcount()):
        neighbors = g.neighbors(node)
        if len(neighbors) < 2:
            clustering_coeffs.append(0.0)
            continue
        
        neighbor_connections = 0
        for i in range(len(neighbors)):
            for j in range(i+1, len(neighbors)):
                if g.are_connected(neighbors[i], neighbors[j]):
                    neighbor_connections += 1
        
        possible_connections = len(neighbors) * (len(neighbors) - 1) / 2
        clustering_coeffs.append(neighbor_connections / possible_connections if possible_connections > 0 else 0.0)
    
    return np.mean(clustering_coeffs)


def calculate_all_network_metrics(g, degrees, largest_component, rr_indices):
    """计算所有网络指标"""
    metrics = {}
    
    # 基础指标
    metrics['n_nodes'] = g.vcount()
    metrics['n_edges'] = g.ecount()
    metrics['density'] = g.density()
    metrics['avg_degree'] = np.mean(degrees) if len(degrees) > 0 else 0
    metrics['max_degree'] = np.max(degrees) if len(degrees) > 0 else 0
    metrics['min_degree'] = np.min(degrees) if len(degrees) > 0 else 0
    metrics['clustering_coeff'] = calculate_clustering_coefficient(g)
    
    # 连通性指标
    connected_components = g.components()
    metrics['n_components'] = len(connected_components)
    metrics['largest_component_size'] = largest_component.vcount() if largest_component else 0
    metrics['largest_component_ratio'] = metrics['largest_component_size'] / metrics['n_nodes'] if metrics['n_nodes'] > 0 else 0
    
    # 路径长度指标
    path_metrics = calculate_path_metrics(largest_component)
    metrics.update(path_metrics)
    
    # 效率指标
    efficiency_metrics = calculate_efficiency_metrics(largest_component)
    metrics.update(efficiency_metrics)
    
    # 中心性指标
    centrality_metrics = calculate_centrality_metrics(g, degrees, largest_component, rr_indices)
    metrics.update(centrality_metrics)
    
    # 模块化指标
    modularity_metrics = calculate_modularity_metrics(g)
    metrics.update(modularity_metrics)
    
    # 小世界性
    small_world_metrics = calculate_small_world_metrics(g)
    metrics.update(small_world_metrics)
    
    # 同配性
    assortativity_metrics = calculate_assortativity_metrics(g)
    metrics.update(assortativity_metrics)
    
    # 富俱乐部系数
    rich_club_metrics = calculate_rich_club_metrics_manual(g, degrees)
    metrics.update(rich_club_metrics)
    
    return metrics


def calculate_path_metrics(largest_component):
    """计算路径相关指标"""
    metrics = {}
    
    if largest_component and largest_component.vcount() > 1:
        try:
            # 平均最短路径长度
            avg_path_length = np.mean(largest_component.shortest_paths())
            metrics['avg_path_length'] = avg_path_length
            
            # 网络直径
            diameter = largest_component.diameter()
            metrics['diameter'] = diameter
            
        except Exception as e:
            log.warning(f"计算路径指标时出错: {e}")
            metrics['avg_path_length'] = float('inf')
            metrics['diameter'] = 0
    else:
        metrics['avg_path_length'] = float('inf')
        metrics['diameter'] = 0
    
    return metrics


def calculate_efficiency_metrics(largest_component):
    """计算效率相关指标"""
    metrics = {}
    
    if largest_component and largest_component.vcount() > 1:
        try:
            # 全局效率
            distances = largest_component.shortest_paths()
            efficiencies = []
            for i in range(len(distances)):
                for j in range(i+1, len(distances)):
                    if distances[i][j] != float('inf') and distances[i][j] > 0:
                        efficiencies.append(1.0 / distances[i][j])
            
            metrics['global_efficiency'] = np.mean(efficiencies) if efficiencies else 0.0
            
            # 局部效率 (简化版本)
            local_efficiencies = []
            for node in range(largest_component.vcount()):
                neighbors = largest_component.neighbors(node)
                if len(neighbors) > 1:
                    subgraph = largest_component.induced_subgraph(neighbors)
                    if subgraph.vcount() > 1:
                        sub_distances = subgraph.shortest_paths()
                        sub_efficiencies = []
                        for i in range(len(sub_distances)):
                            for j in range(i+1, len(sub_distances)):
                                if sub_distances[i][j] != float('inf') and sub_distances[i][j] > 0:
                                    sub_efficiencies.append(1.0 / sub_distances[i][j])
                        if sub_efficiencies:
                            local_efficiencies.append(np.mean(sub_efficiencies))
            
            metrics['local_efficiency'] = np.mean(local_efficiencies) if local_efficiencies else 0.0
            
        except Exception as e:
            log.warning(f"计算效率指标时出错: {e}")
            metrics['global_efficiency'] = 0.0
            metrics['local_efficiency'] = 0.0
    else:
        metrics['global_efficiency'] = 0.0
        metrics['local_efficiency'] = 0.0
    
    return metrics


def calculate_centrality_metrics(g, degrees, largest_component, rr_indices):
    """计算中心性相关指标"""
    metrics = {}
    
    if g.vcount() == 0:
        return _get_empty_centrality_metrics()
    
    try:
        # 度中心性
        degree_centrality = np.array(degrees) / (g.vcount() - 1) if g.vcount() > 1 else np.zeros_like(degrees)
        metrics['avg_degree_centrality'] = float(np.mean(degree_centrality)) if len(degree_centrality) > 0 else 0.0
        
        # 介数中心性
        try:
            betweenness = g.betweenness()
            if isinstance(betweenness, list) and len(betweenness) > 0:
                betweenness_array = np.array(betweenness)
                metrics['avg_betweenness'] = float(np.mean(betweenness_array))
                metrics['max_betweenness'] = float(np.max(betweenness_array))
            else:
                metrics['avg_betweenness'] = 0.0
                metrics['max_betweenness'] = 0.0
        except Exception as e:
            log.warning(f"计算介数中心性时出错: {e}")
            metrics['avg_betweenness'] = 0.0
            metrics['max_betweenness'] = 0.0
        
        # 紧密度中心性 (在最大连通分量上计算)
        try:
            if largest_component and largest_component.vcount() > 1:
                closeness = largest_component.closeness()
                if isinstance(closeness, list) and len(closeness) > 0:
                    closeness_array = np.array(closeness)
                    metrics['avg_closeness'] = float(np.mean(closeness_array))
                else:
                    metrics['avg_closeness'] = 0.0
            else:
                metrics['avg_closeness'] = 0.0
        except Exception as e:
            log.warning(f"计算紧密度中心性时出错: {e}")
            metrics['avg_closeness'] = 0.0
        
        # Hub分析 - 统一使用z-score > 1.5标准
        if len(degrees) > 1:
            try:
                # 使用统一的z-score阈值判断Hub节点
                z_scores = zscore(degrees)
                hubs = np.where(z_scores > HUB_ZSCORE_THRESHOLD)[0]
                metrics['n_hubs'] = int(len(hubs))
                # 移除了 hub_fraction 指标
                
                # 保存原始数据中的索引和本地索引
                if rr_indices is not None and len(hubs) > 0:
                    metrics['hub_indices_original'] = rr_indices[hubs].tolist()  # 在完整数据集中的原始索引
                    metrics['hub_indices_local'] = hubs.tolist()  # 在RR子集中的本地索引
                else:
                    metrics['hub_indices_original'] = []
                    metrics['hub_indices_local'] = hubs.tolist()
                
                metrics['hub_degrees'] = degrees[hubs].tolist() if len(hubs) > 0 else []
                
                # hub_z_scores保留三位有效数字
                if len(hubs) > 0:
                    formatted_z_scores = [float(f"{score:.3g}") for score in z_scores[hubs]]
                    metrics['hub_z_scores'] = formatted_z_scores
                else:
                    metrics['hub_z_scores'] = []
                    
            except Exception as e:
                log.warning(f"计算Hub指标时出错: {e}")
                metrics.update(_get_empty_hub_metrics())
        else:
            metrics.update(_get_empty_hub_metrics())
                
    except Exception as e:
        log.warning(f"计算中心性指标时出错: {e}")
        metrics.update(_get_empty_centrality_metrics())
    
    return metrics


def _get_empty_centrality_metrics():
    """返回空的中心性指标"""
    return {
        'avg_degree_centrality': 0.0,
        'avg_betweenness': 0.0,
        'max_betweenness': 0.0,
        'avg_closeness': 0.0
    }


def _get_empty_hub_metrics():
    """返回空的Hub指标"""
    return {
        'n_hubs': 0,
        'hub_indices_original': [],
        'hub_indices_local': [],
        'hub_degrees': [],
        'hub_z_scores': [],
    }


def calculate_modularity_metrics(g):
    """计算模块化相关指标"""
    metrics = {}
    
    if g.vcount() > 1 and g.ecount() > 0:
        try:
            # 使用Louvain算法检测社区
            communities = g.community_multilevel()
            modularity = g.modularity(communities)
            metrics['modularity'] = modularity
            metrics['n_communities'] = len(communities)
            metrics['avg_community_size'] = np.mean([len(c) for c in communities])
            
        except Exception as e:
            log.warning(f"计算模块化指标时出错: {e}")
            metrics['modularity'] = 0
            metrics['n_communities'] = 1
            metrics['avg_community_size'] = g.vcount()
    else:
        metrics['modularity'] = 0
        metrics['n_communities'] = 1
        metrics['avg_community_size'] = g.vcount() if g.vcount() > 0 else 0
    
    return metrics


def calculate_small_world_metrics(g):
    """计算小世界性相关指标"""
    metrics = {}
    
    if g.vcount() < 10:  # 网络太小无法可靠计算
        metrics['small_worldness'] = 0.0
        return metrics
    
    try:
        # 计算实际网络的聚类系数和平均路径长度
        C_real = calculate_clustering_coefficient(g)
        
        # 计算实际网络的平均路径长度（使用最大连通分量）
        giant = g.components().giant()
        if giant.vcount() < 2:
            metrics['small_worldness'] = 0.0
            return metrics
        
        L_real = np.mean(giant.shortest_paths())
        
        # 生成随机网络并计算平均值
        n_random = 3  # 减少随机网络数量以提高速度
        C_random_list = []
        L_random_list = []
        
        for _ in range(n_random):
            # 生成相同节点数和边数的随机网络
            random_g = ig.Graph.Erdos_Renyi(n=g.vcount(), m=g.ecount())
            C_random_list.append(calculate_clustering_coefficient(random_g))
            
            # 计算随机网络的平均路径长度
            random_giant = random_g.components().giant()
            if random_giant.vcount() > 1:
                try:
                    L_random = np.mean(random_giant.shortest_paths())
                    L_random_list.append(L_random)
                except:
                    pass
        
        C_random = np.mean(C_random_list) if C_random_list else 1.0
        L_random = np.mean(L_random_list) if L_random_list else 1.0
        
        # 小世界性 = (C_real / C_random) / (L_real / L_random)
        if C_random > 0 and L_random > 0 and L_real > 0:
            small_worldness = (C_real / C_random) / (L_real / L_random)
            metrics['small_worldness'] = small_worldness
        else:
            metrics['small_worldness'] = 0.0
            
    except Exception as e:
        log.warning(f"计算小世界性时出错: {e}")
        metrics['small_worldness'] = 0.0
    
    return metrics


def calculate_assortativity_metrics(g):
    """计算同配性相关指标"""
    metrics = {}
    
    if g.vcount() > 1 and g.ecount() > 0:
        try:
            # 度同配性
            assortativity = g.assortativity_degree()
            metrics['assortativity'] = assortativity
            
        except Exception as e:
            log.warning(f"计算同配性时出错: {e}")
            metrics['assortativity'] = 0.0
    else:
        metrics['assortativity'] = 0.0
    
    return metrics


def calculate_rich_club_metrics_manual(g, degrees):
    """手动计算富俱乐部系数"""
    metrics = {}
    
    if g.vcount() < 3:
        metrics['avg_rich_club'] = 0.0
        metrics['max_rich_club'] = 0.0
        return metrics
    
    try:
        # 富俱乐部系数定义：对于度数为k的节点，计算这些高度数节点之间实际边数与可能最大边数的比例
        max_degree = int(np.max(degrees))
        
        # 只计算到最大度数的一半，避免统计不可靠
        max_k = min(max_degree // 2, 10)
        
        if max_k < 1:
            metrics['avg_rich_club'] = 0.0
            metrics['max_rich_club'] = 0.0
            return metrics
        
        rich_club_coeffs = []
        
        for k in range(1, max_k + 1):
            # 找出度数大于k的节点
            high_degree_nodes = [i for i, deg in enumerate(degrees) if deg > k]
            n_high = len(high_degree_nodes)
            
            if n_high < 2:
                continue
            
            # 计算这些高度数节点之间的实际边数
            actual_edges = 0
            for i in range(n_high):
                for j in range(i + 1, n_high):
                    if g.are_connected(high_degree_nodes[i], high_degree_nodes[j]):
                        actual_edges += 1
            
            # 可能的边数
            possible_edges = n_high * (n_high - 1) / 2
            
            if possible_edges > 0:
                rich_club_coeff = actual_edges / possible_edges
                rich_club_coeffs.append(rich_club_coeff)
        
        if rich_club_coeffs:
            metrics['avg_rich_club'] = float(np.mean(rich_club_coeffs))
            metrics['max_rich_club'] = float(np.max(rich_club_coeffs))
        else:
            metrics['avg_rich_club'] = 0.0
            metrics['max_rich_club'] = 0.0
            
    except Exception as e:
        log.warning(f"计算富俱乐部系数时出错: {e}")
        metrics['avg_rich_club'] = 0.0
        metrics['max_rich_club'] = 0.0
    
    return metrics


def print_network_metrics(metrics, rr_categories, threshold, stimulus_type="All"):
    """打印所有网络指标"""
    log.info("\n" + "="*70)
    log.info(f"            RR神经元网络分析结果 (刺激类型: {stimulus_type}, 阈值={threshold})")
    log.info("="*70)
    
    # 基础信息
    log.info(f"\n📊 基础信息:")
    log.info(f"   • 神经元总数: {metrics['n_nodes']}")
    log.info(f"   • 兴奋性神经元: {np.sum(rr_categories == 'exc')}")
    log.info(f"   • 抑制性神经元: {np.sum(rr_categories == 'inh')}")
    
    # 网络结构指标
    log.info(f"\n🏗️  网络结构:")
    log.info(f"   • 边数: {metrics['n_edges']}")
    log.info(f"   • 网络密度: {metrics['density']:.3g}")
    log.info(f"   • 平均度数: {metrics['avg_degree']:.3g}")
    log.info(f"   • 最大度数: {metrics['max_degree']}")
    log.info(f"   • 连通分量数: {metrics['n_components']}")
    log.info(f"   • 最大连通分量比例: {metrics['largest_component_ratio']:.3g}")
    
    # 聚类和路径指标
    log.info(f"\n🔗 拓扑特性:")
    log.info(f"   • 平均聚类系数: {metrics['clustering_coeff']:.3g}")
    log.info(f"   • 平均最短路径: {metrics['avg_path_length']:.3g}")
    log.info(f"   • 网络直径: {metrics['diameter']}")
    
    # 效率指标
    log.info(f"\n⚡ 效率指标:")
    log.info(f"   • 全局效率: {metrics['global_efficiency']:.3g}")
    log.info(f"   • 局部效率: {metrics['local_efficiency']:.3g}")
    
    # 中心性指标
    log.info(f"\n🎯 中心性分析:")
    log.info(f"   • 平均度中心性: {metrics['avg_degree_centrality']:.3g}")
    log.info(f"   • 平均介数中心性: {metrics['avg_betweenness']:.3g}")
    log.info(f"   • 平均紧密度中心性: {metrics['avg_closeness']:.3g}")
    
    # Hub分析 - 移除了hub_fraction
    log.info(f"\n🎯 Hub节点分析 (z-score > {HUB_ZSCORE_THRESHOLD}):")
    log.info(f"   • Hub节点数量: {metrics['n_hubs']}个")
    
    if metrics['n_hubs'] > 0:
        # 显示原始索引（在完整数据集中的索引）
        if len(metrics['hub_indices_original']) > 0:
            log.info(f"   • Hub节点原始索引: {metrics['hub_indices_original']}")
        # 显示本地索引（在RR子集中的索引）
        if len(metrics['hub_indices_local']) > 0:
            log.info(f"   • Hub节点本地索引: {metrics['hub_indices_local']}")
        
        if len(metrics['hub_degrees']) > 0:
            log.info(f"   • Hub节点度数范围: {np.min(metrics['hub_degrees'])} - {np.max(metrics['hub_degrees'])}")
        if len(metrics['hub_z_scores']) > 0:
            min_z = float(f"{np.min(metrics['hub_z_scores']):.3g}")
            max_z = float(f"{np.max(metrics['hub_z_scores']):.3g}")
            log.info(f"   • Hub节点z-score范围: {min_z:.3g} - {max_z:.3g}")
    
    # 模块化指标
    log.info(f"\n🧩 模块化分析:")
    log.info(f"   • 模块度: {metrics['modularity']:.3g}")
    log.info(f"   • 社区数量: {metrics['n_communities']}")
    log.info(f"   • 平均社区大小: {metrics['avg_community_size']:.3g}")
    
    # 高级网络特性
    log.info(f"\n🌟 高级网络特性:")
    log.info(f"   • 小世界性: {metrics['small_worldness']:.3g}")
    log.info(f"   • 同配性: {metrics['assortativity']:.3g}")
    log.info(f"   • 平均富俱乐部系数: {metrics['avg_rich_club']:.3g}")
    log.info(f"   • 最大富俱乐部系数: {metrics['max_rich_club']:.3g}")
    
    log.info("="*70)


# -------------------- 阈值扫描功能 --------------------
def threshold_scan_analysis(fluorescence_rr, rr_categories, rr_indices, stimulus_type="All"):
    """执行皮尔逊系数阈值扫描分析"""
    log.info("\n" + "="*70)
    log.info(f"             开始皮尔逊系数阈值扫描分析 (刺激类型: {stimulus_type})")
    log.info("="*70)
    
    # 使用配置的阈值范围
    thresholds = SCAN_THRESHOLDS
    all_metrics = []
    
    # 计算相关性矩阵（只需计算一次）
    log.info("计算相关性矩阵...")
    correlation_matrix = calculate_correlation_matrix_rr(fluorescence_rr)
    
    for threshold in thresholds:
        log.info(f"\n>>> 正在分析阈值: {threshold:.2f}")
        
        try:
            # 网络分析（传递rr_indices用于Hub索引映射）
            g, degrees, network_metrics = create_network_analysis(correlation_matrix, threshold, rr_indices)
            
            # 添加阈值信息
            network_metrics['threshold'] = threshold
            network_metrics['stimulus_type'] = stimulus_type
            
            # 添加神经元类型信息
            network_metrics['n_exc_neurons'] = np.sum(rr_categories == 'exc')
            network_metrics['n_inh_neurons'] = np.sum(rr_categories == 'inh')
            
            all_metrics.append(network_metrics)
            
            # 打印当前阈值的结果摘要
            log.info(f"   节点数: {network_metrics['n_nodes']}, 边数: {network_metrics['n_edges']}, "
                   f"密度: {network_metrics['density']:.3g}, Hub节点: {network_metrics['n_hubs']}个")
            
        except Exception as e:
            log.error(f"阈值 {threshold} 分析失败: {e}")
            continue
    
    # 保存结果到DataFrame但不保存文件
    if all_metrics:
        df = pd.DataFrame(all_metrics)
        
        # 重新排列列的顺序，让threshold和stimulus_type在前
        cols = ['threshold', 'stimulus_type'] + [col for col in df.columns if col not in ['threshold', 'stimulus_type']]
        df = df[cols]
        
        log.info(f"\n✅ {stimulus_type}阈值扫描完成！")
        log.info(f"   共分析了 {len(all_metrics)} 个阈值")
        
        # 显示一些统计信息
        log.info(f"\n📈 {stimulus_type}阈值扫描统计摘要:")
        log.info(f"   阈值范围: {thresholds[0]:.2f} - {thresholds[-1]:.2f}")
        log.info(f"   边数范围: {df['n_edges'].min()} - {df['n_edges'].max()}")
        log.info(f"   密度范围: {df['density'].min():.3g} - {df['density'].max():.3g}")
        log.info(f"   Hub节点数量范围: {df['n_hubs'].min()} - {df['n_hubs'].max()}")
        
        return df
    else:
        log.error(f"❌ {stimulus_type}没有成功分析任何阈值")
        return None


# -------------------- 对比分析功能 (修改为支持8个时间段) --------------------
def compare_stimulus_periods(all_results, output_dir):
    """对比8个时间段（IC、LC、IC2、IC4、LC2、LC4、基线、空白屏幕）的网络指标，包含所有阈值结果"""
    log.info("\n" + "="*70)
    log.info("             开始8个时间段的网络指标对比 (包含所有阈值)")
    log.info("="*70)
    
    if not all_results:
        log.warning("❌ 没有可对比的结果数据")
        return
    
    # 收集所有时间段的所有阈值结果
    comparison_data = []
    
    for stimulus_type, results in all_results.items():
        if results is not None:
            # 包含所有阈值的结果
            for _, row in results.iterrows():
                metrics = row.to_dict()
                metrics['stimulus_type'] = stimulus_type
                comparison_data.append(metrics)
    
    if len(comparison_data) < 2:
        log.warning("❌ 可对比的数据数量不足")
        return
    
    # 创建对比表格
    df_comparison = pd.DataFrame(comparison_data)
    
    # 确保阈值列在最前面
    if 'threshold' in df_comparison.columns:
        cols = ['threshold', 'stimulus_type'] + [col for col in df_comparison.columns if col not in ['threshold', 'stimulus_type']]
        df_comparison = df_comparison[cols]
    
    # 选择关键指标进行对比（移除了hub_fraction）
    key_metrics = [
        'n_nodes', 'n_edges', 'density', 'avg_degree', 'clustering_coeff',
        'avg_path_length', 'global_efficiency', 'local_efficiency',
        'modularity', 'small_worldness', 'assortativity', 'n_hubs'
    ]
    
    # 过滤出存在的指标
    available_metrics = [metric for metric in key_metrics if metric in df_comparison.columns]
    
    # 保存完整的对比结果（包含所有阈值）
    comparison_csv = os.path.join(output_dir, "network_comparison_eight_periods_all_thresholds.csv")
    df_comparison.to_csv(comparison_csv, index=False, encoding='utf-8-sig')
    
    log.info(f"\n📊 8个时间段网络指标对比 (所有阈值):")
    
    # 定义时间段显示顺序
    stimulus_order = ['IC', 'LC', 'IC2', 'IC4', 'LC2', 'LC4', 'Baseline', 'Blank_Screen']
    
    for stimulus_type in stimulus_order:
        if stimulus_type in df_comparison['stimulus_type'].unique():
            subset = df_comparison[df_comparison['stimulus_type'] == stimulus_type]
            log.info(f"\n🎯 {stimulus_type}刺激期间 (阈值范围: {subset['threshold'].min():.2f}-{subset['threshold'].max():.2f}):")
            log.info(f"   节点数: {subset['n_nodes'].mean():.1f}±{subset['n_nodes'].std():.1f}")
            log.info(f"   边数: {subset['n_edges'].mean():.1f}±{subset['n_edges'].std():.1f}")
            log.info(f"   密度: {subset['density'].mean():.3g}±{subset['density'].std():.3g}")
            log.info(f"   聚类系数: {subset['clustering_coeff'].mean():.3g}±{subset['clustering_coeff'].std():.3g}")
            log.info(f"   Hub节点: {subset['n_hubs'].mean():.1f}±{subset['n_hubs'].std():.1f}个")

    # 自动转换为Excel格式（保留三位有效数字）
    comparison_excel = os.path.join(output_dir, "network_comparison_eight_periods_all_thresholds_formatted.xlsx")
    csv_to_excel(comparison_csv, comparison_excel)
    
    log.info(f"\n✅ 8个时间段对比分析完成！")
    log.info(f"   • 完整对比表格CSV: {comparison_csv}")
    log.info(f"   • 完整对比表格Excel: {comparison_excel}")
    
    return df_comparison


# -------------------- 分刺激类型分析 (修改为支持8个时间段) --------------------
def analyze_by_stimulus_type(fluorescence_rr, rr_categories, rr_indices, output_dir, data_path):
    """按刺激类型分别进行分析（IC、LC、IC2、IC4、LC2、LC4、基线、空白屏幕），只生成对比结果文件"""
    log.info("\n" + "="*70)
    log.info("             开始按刺激类型分别分析 (IC、LC、IC2、IC4、LC2、LC4、基线、空白屏幕)")
    log.info("="*70)
    
    # 使用RR分析保存的时间段信息分割数据（8个时间段）
    ic_data, lc_data, ic2_data, ic4_data, lc2_data, lc4_data, baseline_data, blank_screen_data = split_data_into_eight_periods(fluorescence_rr, data_path)
    
    all_results = {}
    
    # 分析8个时间段
    periods = [
        ("IC", ic_data),
        ("LC", lc_data),
        ("IC2", ic2_data),
        ("IC4", ic4_data),
        ("LC2", lc2_data), 
        ("LC4", lc4_data),
        ("Baseline", baseline_data),
        ("Blank_Screen", blank_screen_data)
    ]
    
    for stimulus_type, data in periods:
        if data.shape[1] > 0:
            log.info(f"\n🔊 分析{stimulus_type}期间数据 (时间点: {data.shape[1]})")
            results = threshold_scan_analysis(data, rr_categories, rr_indices, stimulus_type)
            all_results[stimulus_type] = results
        else:
            log.warning(f"❌ {stimulus_type}期间数据为空，跳过分析")
    
    # 执行对比分析
    comparison_results = compare_stimulus_periods(all_results, output_dir)
    
    return all_results, comparison_results


# -------------------- 主流程 (修改为支持8个时间段) --------------------
def main():
    try:
        log.info("=== RR 神经元网络分析开始 ===")
        
        # 显示当前模式配置
        log.info(f"🐭 分析小鼠: {MOUSE_ID}")
        if USE_EXCITATORY_ONLY:
            log.info(f"🎯 神经元选择: 仅兴奋性RR神经元")
        else:
            log.info(f"🎯 神经元选择: 全部RR神经元 (兴奋性+抑制性)")
        
        log.info(f"📊 分析模式: 阈值扫描 (范围: {SCAN_THRESHOLDS[0]:.2f} - {SCAN_THRESHOLDS[-1]:.2f})")
        log.info(f"🎯 Hub节点判断标准: z-score > {HUB_ZSCORE_THRESHOLD}")
        log.info(f"🔊 刺激类型分析: 开启 (IC、LC、IC2、IC4、LC2、LC4、基线、空白屏幕)")
        log.info(f"📈 高通滤波: 截止频率 {HIGH_PASS_CUTOFF} Hz, 采样率 {SAMPLING_RATE} Hz")
        log.info(f"📊 数值精度: 保留三位有效数字")
        
        # 加载数据（应用高通滤波）
        fluorescence_rr, coordinates_rr, rr_indices, rr_categories = load_rr_neurons_data(DATA_FILE, RR_INDICES_CSV, apply_filter=True)
        
        # 确定输出目录和数据目录
        output_dir = os.path.dirname(DATA_FILE)
        data_path = output_dir  # 数据目录与输出目录相同
        
        # 检查是否存在RR分析的时间段文件
        required_files = [
            "stimulus_periods_ic2.npy",
            "stimulus_periods_ic4.npy",
            "stimulus_periods_lc2.npy",
            "stimulus_periods_lc4.npy",
            "stimulus_periods_baseline.npy",
            "stimulus_periods_blank_screen.npy"
        ]
        
        existing_files = [f for f in required_files if os.path.exists(os.path.join(data_path, f))]
        
        if len(existing_files) == len(required_files):
            log.info("✅ 检测到完整的RR分析时间段文件，将使用精确的时间段分割")
        elif len(existing_files) > 0:
            log.warning(f"⚠️  找到部分时间段文件 ({len(existing_files)}/{len(required_files)})，将使用可用时间段")
        else:
            log.warning("⚠️  未找到RR分析时间段文件，将使用默认分割方法")
        
        # 按刺激类型分别分析（IC、LC、IC2、IC4、LC2、LC4、基线、空白屏幕）
        all_results, comparison_results = analyze_by_stimulus_type(fluorescence_rr, rr_categories, rr_indices, output_dir, data_path)
        
        log.info(f"\n📊 所有刺激类型分析完成!")
        for stimulus_type, results in all_results.items():
            if results is not None:
                log.info(f"   • {stimulus_type}: 分析完成")
        
        log.info("\n🎉 所有分析完成!")

    except Exception as e:
        log.error(f"处理过程中出现错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()